#!/usr/bin/env python3
"""
配布員給与データをExcelテンプレートに差し込むスクリプト

Usage: python3 fill-payroll-excel.py <input_xlsx> <password> <payroll_json> <output_xlsx>
"""

import sys
import json
import io
import msoffcrypto
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from copy import copy

ALERT_THRESHOLD_RATE = 0.2
ALERT_THRESHOLD_ABS = 2000
ALERT_FILL = PatternFill(start_color="FFFECACA", end_color="FFFECACA", fill_type="solid")
NEW_STAFF_FILL = PatternFill(start_color="FFE0F2FE", end_color="FFE0F2FE", fill_type="solid")
ZERO_FILL = PatternFill(start_color="FFFF99CC", end_color="FFFF99CC", fill_type="solid")      # pink - 7日間0円
HAS_DATA_FILL = PatternFill(start_color="FF99CCFF", end_color="FF99CCFF", fill_type="solid")  # blue - 金額あり
HEADER_FONT = Font(bold=True, size=9)


def decrypt_workbook(path, password):
    with open(path, "rb") as f:
        buf = io.BytesIO()
        ms = msoffcrypto.OfficeFile(f)
        ms.load_key(password=password)
        ms.decrypt(buf)
    return openpyxl.load_workbook(buf)


def find_target_sheet(wb, week_start_date):
    year = week_start_date.year
    month = week_start_date.month
    candidates = []
    for name in wb.sheetnames:
        if f"{year}年{month}月" in name:
            candidates.append(name)
        week_end = week_start_date + timedelta(days=6)
        if f"{year}年{week_end.month}月" in name and name not in candidates:
            candidates.append(name)
    for c in candidates:
        if "一般" in c:
            return c
    return candidates[0] if candidates else None


def find_week_block(ws, week_start_date):
    target_date = week_start_date.date()
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val is None:
            continue
        if hasattr(val, 'date'):
            cell_date = val.date()
        elif hasattr(val, 'year'):
            cell_date = val
        else:
            continue
        if cell_date == target_date:
            return r
    return None


def find_staff_columns(ws, staff_code_row=20):
    staff_map = {}
    for c in range(2, ws.max_column + 1):
        val = ws.cell(staff_code_row, c).value
        if val and str(val).strip():
            staff_map[str(val).strip()] = c
    return staff_map


def find_insert_position(staff_cols, new_staff_id):
    """スタッフコード順でソートされた位置に挿入する列を返す"""
    # 既存のMBF/MSF/MYFスタッフコードのみ対象
    existing = sorted(
        [(sid, col) for sid, col in staff_cols.items()],
        key=lambda x: x[1]  # 列番号順
    )

    # 新しいスタッフIDが入るべき位置を探す
    # スタッフコードの大小比較で、直後に来るべき列を見つける
    insert_after_col = None
    for sid, col in existing:
        if sid < new_staff_id:
            insert_after_col = col
        else:
            break

    if insert_after_col is not None:
        return insert_after_col + 1
    # 全員より前に来る場合は最初のスタッフ列
    if existing:
        return existing[0][1]
    return 2


def insert_column(ws, col_idx):
    """指定位置に空列を挿入する（openpyxlのinsert_cols）"""
    ws.insert_cols(col_idx)


def copy_column_style(ws, source_col, target_col, max_row):
    """ソース列のスタイルをターゲット列にコピー"""
    for r in range(1, max_row + 1):
        src = ws.cell(r, source_col)
        tgt = ws.cell(r, target_col)
        if src.has_style:
            tgt.font = copy(src.font)
            tgt.border = copy(src.border)
            tgt.number_format = src.number_format
            tgt.alignment = copy(src.alignment)
            # fill はコピーしない（後で独自に設定する）


def main():
    if len(sys.argv) != 5:
        print("Usage: fill-payroll-excel.py <input> <password> <payroll_json> <output>", file=sys.stderr)
        sys.exit(1)

    input_path, password, json_path, output_path = sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4]

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    week_start = datetime.strptime(data["weekStart"], "%Y-%m-%d")
    distributors = data["distributors"]

    wb = decrypt_workbook(input_path, password)

    sheet_name = find_target_sheet(wb, week_start)
    if not sheet_name:
        print(json.dumps({"error": f"シートが見つかりません: {week_start.year}年{week_start.month}月"}))
        sys.exit(0)

    ws = wb[sheet_name]

    block_start = find_week_block(ws, week_start)
    if not block_start:
        print(json.dumps({"error": f"週ブロックが見つかりません: {data['weekStart']}"}))
        sys.exit(0)

    subtotal_row = block_start + 7
    expense_row = block_start + 11
    total_row = block_start + 22

    staff_cols = find_staff_columns(ws)

    alerts = []
    new_staff = []
    updated_count = 0

    week_dates = [(week_start + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]

    # --- Phase 1: 新規スタッフの列を挿入（先にやらないと列がずれる） ---
    new_staff_ids = [d["staffId"] for d in distributors if d["staffId"] not in staff_cols]
    # スタッフコード順にソートして挿入
    new_staff_ids.sort()

    for new_id in new_staff_ids:
        # 挿入位置を決定（現在のstaff_colsに基づく）
        insert_col = find_insert_position(staff_cols, new_id)

        # 列を挿入
        insert_column(ws, insert_col)

        # 挿入によって既存の列番号がずれるので更新
        updated_staff_cols = {}
        for sid, col in staff_cols.items():
            if col >= insert_col:
                updated_staff_cols[sid] = col + 1
            else:
                updated_staff_cols[sid] = col
        updated_staff_cols[new_id] = insert_col
        staff_cols = updated_staff_cols

        # 前の列のスタイルをコピー（見た目を揃える）
        style_src = insert_col - 1 if insert_col > 1 else insert_col + 1
        copy_column_style(ws, style_src, insert_col, ws.max_row)

        # ヘッダー書き込み
        dist_data = next((d for d in distributors if d["staffId"] == new_id), None)
        ws.cell(20, insert_col).value = new_id
        ws.cell(20, insert_col).font = HEADER_FONT
        ws.cell(20, insert_col).fill = NEW_STAFF_FILL
        if dist_data:
            ws.cell(21, insert_col).value = dist_data["name"]
            ws.cell(21, insert_col).font = HEADER_FONT
            ws.cell(21, insert_col).fill = NEW_STAFF_FILL

        new_staff.append(new_id)

    # --- Phase 2: データ差し込み + セル色設定 ---
    for dist in distributors:
        staff_id = dist["staffId"]
        col = staff_cols.get(staff_id)
        if col is None:
            continue  # should not happen after Phase 1

        daily_earnings = dist.get("dailyEarnings", {})
        daily_expenses = dist.get("dailyExpenses", {})
        schedule_pay = dist.get("schedulePay", 0)
        expense_pay = dist.get("expensePay", 0)
        gross_pay = dist.get("grossPay", 0)

        # 7日間の合計が0かどうか判定
        total_week_earnings = sum(daily_earnings.get(d, 0) for d in week_dates)
        is_zero_week = total_week_earnings == 0

        # 日別セルの処理
        for day_idx, date_str in enumerate(week_dates):
            r = block_start + day_idx
            new_val = daily_earnings.get(date_str, 0)

            old_val = ws.cell(r, col).value
            old_num = 0
            if old_val is not None:
                try:
                    old_num = float(old_val)
                except (ValueError, TypeError):
                    old_num = 0

            # セル色設定（金額の有無に関わらず）
            if is_zero_week:
                ws.cell(r, col).fill = ZERO_FILL       # ピンク: 7日間0円
            else:
                ws.cell(r, col).fill = HAS_DATA_FILL    # 水色: 金額あり

            # 金額差異アラート
            if old_num > 0 and new_val > 0:
                diff = abs(new_val - old_num)
                rate = diff / old_num if old_num != 0 else 0
                if diff >= ALERT_THRESHOLD_ABS or rate >= ALERT_THRESHOLD_RATE:
                    alerts.append({
                        "staffId": staff_id,
                        "name": dist["name"],
                        "date": date_str,
                        "old": int(old_num),
                        "new": int(new_val),
                        "diff": int(new_val - old_num),
                    })
                    ws.cell(r, col).fill = ALERT_FILL   # 赤: アラートが優先

            # 値の書き込み（0でも書き込む）
            ws.cell(r, col).value = new_val

        # 小計
        if schedule_pay >= 0:
            old_sub = ws.cell(subtotal_row, col).value
            old_sub_num = 0
            if old_sub is not None:
                try:
                    old_sub_num = float(old_sub)
                except (ValueError, TypeError):
                    old_sub_num = 0

            if old_sub_num > 0 and schedule_pay > 0:
                diff = abs(schedule_pay - old_sub_num)
                rate = diff / old_sub_num if old_sub_num != 0 else 0
                if diff >= ALERT_THRESHOLD_ABS or rate >= ALERT_THRESHOLD_RATE:
                    alerts.append({
                        "staffId": staff_id,
                        "name": dist["name"],
                        "date": "小計",
                        "old": int(old_sub_num),
                        "new": int(schedule_pay),
                        "diff": int(schedule_pay - old_sub_num),
                    })
                    ws.cell(subtotal_row, col).fill = ALERT_FILL

            ws.cell(subtotal_row, col).value = schedule_pay

        # 交通費
        ws.cell(expense_row, col).value = expense_pay

        # 合計
        ws.cell(total_row, col).value = gross_pay

        updated_count += 1

    wb.save(output_path)

    result = {
        "success": True,
        "sheet": sheet_name,
        "weekBlock": f"Row {block_start}-{block_start + 22}",
        "updated": updated_count,
        "newStaff": new_staff,
        "alerts": alerts,
    }
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
